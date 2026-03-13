"""
Script to export paid users and their conversations to Excel in hierarchy mode.

Layout:
  - One "user header" row per paid user:
      User ID | User Name | User Created At | Transaction 1 | Transaction 2 | ...
  - Followed by one "conversation" row per conversation of that user (indented):
      (blank) | (blank) | (blank) | Coach ID | Coach Name | Started At | Last Msg At | Messages

Users are sorted by their earliest payment date.
Conversations per user are sorted oldest-first.

Usage:
    python paid_users_excel.py

Requirements:
    - BACKEND_BASE_URL and BACKEND_API_KEY must be configured in .env
    - openpyxl package must be installed

Output:
    paid_users_YYYY-MM-DD_HHMMSS.xlsx
"""
import asyncio
import logging
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from services.backend_client import backend_client

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================
# DATE RANGE FILTER (conversations)
# Format: "YYYY-MM-DD HH:MM:SS" or None to include all
START_DATETIME = None
END_DATETIME = None
# ============================

IST_OFFSET = timedelta(hours=5, minutes=30)

# Light blue fill for user header rows
USER_ROW_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
# Light grey fill for conversation rows
CONV_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def parse_datetime(dt_value) -> Optional[datetime]:
    """Parse a datetime string or object into a timezone-aware datetime."""
    if not dt_value:
        return None
    if isinstance(dt_value, datetime):
        return dt_value if dt_value.tzinfo else dt_value.replace(tzinfo=timezone.utc)
    if isinstance(dt_value, str):
        try:
            return datetime.fromisoformat(dt_value.replace('Z', '+00:00'))
        except ValueError:
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y-%m-%dT%H:%M:%S']:
                try:
                    return datetime.strptime(dt_value, fmt).replace(tzinfo=timezone.utc)
                except ValueError:
                    continue
    return None


def fmt_ist(dt_value) -> str:
    """Convert a UTC datetime to a formatted IST string."""
    dt = parse_datetime(dt_value)
    if not dt:
        return ""
    if not dt.tzinfo:
        dt = dt.replace(tzinfo=timezone.utc)
    return (dt + IST_OFFSET).strftime('%Y-%m-%d %H:%M:%S IST')


def is_in_date_range(conv: Dict, start_dt: Optional[str], end_dt: Optional[str]) -> bool:
    """Return True if conv.startedAt falls within the given range (or range is None)."""
    if not start_dt and not end_dt:
        return True
    conv_dt = parse_datetime(conv.get('startedAt'))
    if not conv_dt:
        return False
    if start_dt:
        s = parse_datetime(start_dt)
        if s:
            if not s.tzinfo:
                s = s.replace(tzinfo=timezone.utc)
            if conv_dt < s:
                return False
    if end_dt:
        e = parse_datetime(end_dt)
        if e:
            if not e.tzinfo:
                e = e.replace(tzinfo=timezone.utc)
            if conv_dt > e:
                return False
    return True


def format_messages(messages: List[Dict], conversation_id: str) -> str:
    """Format all messages for a conversation as 'coach: ...' / 'user: ...' lines."""
    conv_msgs = [m for m in messages if str(m.get('conversationId')) == str(conversation_id)]
    conv_msgs.sort(key=lambda x: x.get('createdAt', x.get('id', '')))
    lines = []
    for msg in conv_msgs:
        text = msg.get('text', '')
        if not text:
            continue
        sender_type = msg.get('senderType', '').upper()
        prefix = 'coach' if sender_type in ('AI_COACH', 'HUMAN_COACH') else 'user'
        lines.append(f"{prefix}: {text}")
    return '\n'.join(lines)


def apply_row_fill(ws, row_idx: int, num_cols: int, fill: PatternFill):
    """Apply a fill color to all cells in a row."""
    for col in range(1, num_cols + 1):
        ws.cell(row=row_idx, column=col).fill = fill


async def export_paid_users_to_excel():
    """Build the hierarchy Excel: user header row(s) + conversation sub-rows."""
    try:
        logger.info("Starting paid users export...")

        if START_DATETIME or END_DATETIME:
            logger.info(f"Date range filter: {START_DATETIME or 'No start'} to {END_DATETIME or 'No end'}")

        logger.info("Fetching all data from backend...")
        res = await backend_client.fetch_all_data()
        data = res.get("data", res)

        conversations = data.get('conversations', [])
        messages = data.get('messages', [])
        users = data.get('users', [])
        razorpay_orders = data.get('razorpayOrders', [])

        logger.info("Fetching coaches from backend...")
        coaches = await backend_client.fetch_coaches()

        logger.info(
            f"Retrieved {len(conversations)} conversations, {len(messages)} messages, "
            f"{len(users)} users, {len(coaches)} coaches, {len(razorpay_orders)} razorpay orders"
        )

        # Lookups
        user_lookup: Dict[str, Dict] = {u['id']: u for u in users if u.get('id')}
        coach_lookup: Dict[str, Dict] = {c['id']: c for c in coaches if c.get('id')}

        # Group paid orders by userId (keep ALL paid orders per user)
        paid_orders_by_user: Dict[str, List[Dict]] = {}
        for order in razorpay_orders:
            if order.get('status') != 'paid':
                continue
            uid = order.get('userId')
            if not uid:
                continue
            paid_orders_by_user.setdefault(uid, []).append(order)

        paid_user_ids = set(paid_orders_by_user.keys())
        logger.info(f"Found {len(paid_user_ids)} users with paid orders")

        # Group conversations by userId (date-range filtered)
        convs_by_user: Dict[str, List[Dict]] = {}
        for conv in conversations:
            uid = conv.get('userId')
            if uid not in paid_user_ids:
                continue
            if not is_in_date_range(conv, START_DATETIME, END_DATETIME):
                continue
            convs_by_user.setdefault(uid, []).append(conv)

        # Sort users by earliest payment date
        sorted_user_ids = sorted(
            paid_user_ids,
            key=lambda uid: min(
                parse_datetime(o.get('createdAt')) or datetime.min.replace(tzinfo=timezone.utc)
                for o in paid_orders_by_user[uid]
            )
        )

        # ----------------------------------------------------------------
        # Column layout
        #
        # USER HEADER ROW (blue):
        #   A: User ID
        #   B: User Name
        #   C: User Created At (IST)
        #   D: Total Paid (INR)
        #   E onwards: one cell per transaction
        #              "Txn N\n₹X | pkg | Y min | <time>"
        #
        # CONVERSATION SUB-ROW (grey, indented):
        #   A: (blank)
        #   B: Coach ID
        #   C: Coach Name
        #   D: Started At (IST)
        #   E: Last Message At (IST)
        #   F: Messages
        # ----------------------------------------------------------------

        # Find the maximum number of transactions any user has
        max_txns = max((len(orders) for orders in paid_orders_by_user.values()), default=1)

        # Fixed columns: A B C D (user info + total), then txn cols, then conv cols
        # Conversation sub-rows use: A(blank) B(coachId) C(coachName) D(startedAt) E(lastMsgAt) F(messages)
        # Total columns needed = max(4 + max_txns, 6)
        total_cols = max(4 + max_txns, 6)

        wb = Workbook()
        ws = wb.active
        ws.title = "Paid Users"

        # Header row
        header = ["User ID", "User Name", "User Registered (IST)", "Total Paid (INR)"]
        for i in range(1, max_txns + 1):
            header.append(f"Transaction {i}")
        # Pad to total_cols if needed (for conv columns)
        while len(header) < total_cols:
            header.append("")
        # Label the conv sub-row columns in the header
        # These are informational — sub-rows align to these columns
        conv_start_col = 1  # conv sub-rows start at col A (blank)
        ws.append(header)
        header_row_cells = ws[1]
        for cell in header_row_cells:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Sub-header row to label conversation columns
        sub_header = ["", "Coach ID", "Coach Name", "Conv. Started (IST)", "Last Message (IST)", "Messages"]
        while len(sub_header) < total_cols:
            sub_header.append("")
        ws.append(sub_header)
        for cell in ws[2]:
            cell.font = Font(bold=True, italic=True)
            cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        logger.info("Writing rows...")
        row_count = 0

        for user_id in sorted_user_ids:
            user_data = user_lookup.get(user_id, {})
            user_name = user_data.get('name', '')
            user_created_at = fmt_ist(user_data.get('createdAt'))

            # Sort this user's payments oldest first
            user_orders = sorted(
                paid_orders_by_user[user_id],
                key=lambda o: parse_datetime(o.get('createdAt')) or datetime.min.replace(tzinfo=timezone.utc)
            )
            total_paid = sum((o.get('amountPaise', 0) or 0) for o in user_orders) / 100

            # Build user header row
            user_row = [user_id, user_name, user_created_at, f"₹{total_paid:.2f}"]
            for order in user_orders:
                amount_inr = (order.get('amountPaise', 0) or 0) / 100
                package_id = order.get('packageId', '') or ''
                minutes = order.get('minutes', '') or ''
                txn_time = fmt_ist(order.get('createdAt'))
                cell_text = f"₹{amount_inr:.2f}\n{package_id}\n{minutes} min\n{txn_time}"
                user_row.append(cell_text)
            while len(user_row) < total_cols:
                user_row.append("")

            ws.append(user_row)
            user_excel_row = ws.max_row
            apply_row_fill(ws, user_excel_row, total_cols, USER_ROW_FILL)
            for col_idx, cell in enumerate(ws[user_excel_row], start=1):
                cell.font = Font(bold=True)
                wrap = col_idx >= 5  # wrap transaction cells
                cell.alignment = Alignment(wrap_text=wrap, vertical='top')

            # Conversation sub-rows for this user
            user_convs = sorted(
                convs_by_user.get(user_id, []),
                key=lambda c: parse_datetime(c.get('startedAt')) or datetime.min.replace(tzinfo=timezone.utc)
            )

            if not user_convs:
                # Placeholder row so there's at least one sub-row
                conv_row = ["", "", "", "No conversations", "", ""]
                while len(conv_row) < total_cols:
                    conv_row.append("")
                ws.append(conv_row)
                apply_row_fill(ws, ws.max_row, total_cols, CONV_ROW_FILL)
            else:
                for conv in user_convs:
                    coach_id = conv.get('coachId', '')
                    coach_data = coach_lookup.get(coach_id, {})
                    first = (coach_data.get('firstName') or '').strip()
                    last = (coach_data.get('lastName') or '').strip()
                    coach_name = f"{first} {last}".strip() or coach_id or ''

                    started_at = fmt_ist(conv.get('startedAt'))
                    last_msg_at = fmt_ist(conv.get('lastMessageAt'))
                    msg_text = format_messages(messages, conv.get('id'))

                    conv_row = ["", coach_id, coach_name, started_at, last_msg_at, msg_text]
                    while len(conv_row) < total_cols:
                        conv_row.append("")

                    ws.append(conv_row)
                    conv_excel_row = ws.max_row
                    apply_row_fill(ws, conv_excel_row, total_cols, CONV_ROW_FILL)
                    for col_idx, cell in enumerate(ws[conv_excel_row], start=1):
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

            row_count += 1
            if row_count % 5 == 0:
                logger.info(f"Processed {row_count} users...")

        # Column widths
        col_widths = {
            'A': 28,   # User ID / blank
            'B': 22,   # User Name / Coach ID
            'C': 26,   # User Created At / Coach Name
            'D': 20,   # Total Paid / Started At
            'E': 26,   # Last Msg At (conv) / Txn 1
            'F': 80,   # Messages (conv) / Txn 2
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
        # Transaction columns beyond F
        for i in range(7, total_cols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 26

        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        filename = f"paid_users_{timestamp}.xlsx"
        wb.save(filename)
        logger.info(f"Successfully exported {row_count} users to {filename}")
        return filename

    except Exception as e:
        logger.error(f"Error exporting paid users: {e}", exc_info=True)
        raise
    finally:
        await backend_client.close()


async def main():
    """Entry point for the script."""
    try:
        filename = await export_paid_users_to_excel()
        print(f"\n{'='*60}")
        print(f"Export completed successfully!")
        print(f"File: {filename}")
        print(f"{'='*60}\n")
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"Export failed: {e}")
        print(f"{'='*60}\n")
        exit(1)


if __name__ == "__main__":
    asyncio.run(main())
