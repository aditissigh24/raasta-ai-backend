"""
Script to export a single conversation to Excel by Conversation ID.

Fetches all data from the backend, filters messages for the specified
conversation ID, and exports in the same format as conversation_excel.py.

Usage:
    python single_conversation_excel.py

Configuration:
    CONVERSATION_ID - the database ID of the conversation to export

Requirements:
    - Backend API must be running and accessible
    - BACKEND_BASE_URL and BACKEND_API_KEY must be configured in .env file
    - openpyxl package must be installed (pip install openpyxl)

Output:
    Creates an Excel file named: conversation_{ID}_{YYYY-MM-DD_HHMMSS}.xlsx

    Columns:
    1. Coach ID        - Database ID of the coach
    2. User ID         - CometChat UID of the user
    3. Conversation ID - CometChat chatroom ID
    4. User Converted  - "Yes" if user converted (isGuest=false), "No" if guest
    5. Messages        - Formatted conversation messages (coach: text\\nuser: text)
    6. Started At      - When the conversation started (IST)
    7. Last Message At - When the last message was sent (IST)
"""
import asyncio
import logging
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from services.backend_client import backend_client

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# IST = UTC + 5:30
IST = timezone(timedelta(hours=5, minutes=30))

# ============================
# CONFIGURATION
# ============================
CONVERSATION_ID = "cmlqv8k6p00yk01sek8fkpspc"   # Set to the conversation DB ID to export
# ============================


def format_messages(messages: List[Dict], conversation_id: str) -> str:
    """Format messages for a conversation into a readable string."""
    conv_messages = [
        m for m in messages if str(m.get('conversationId')) == str(conversation_id)
    ]
    conv_messages.sort(key=lambda x: x.get('createdAt', x.get('id', '')))

    formatted_lines = []
    for msg in conv_messages:
        sender_type = msg.get('senderType', '').upper()
        text = msg.get('text', '')
        if not text:
            continue
        prefix = 'coach' if sender_type in ('AI_COACH', 'HUMAN_COACH') else 'user'
        formatted_lines.append(f"{prefix}: {text}")

    return '\n'.join(formatted_lines)


def get_user_converted_status(user_data: Optional[Dict]) -> str:
    """Return "Yes" if user converted (isGuest=false), "No" otherwise."""
    if not user_data:
        return "Unknown"
    is_guest = user_data.get('isGuest')
    if is_guest is True:
        return "No"
    elif is_guest is False:
        return "Yes"
    return "Unknown"


def format_datetime(dt_value) -> str:
    """Format a UTC datetime value as IST for Excel display."""
    if not dt_value:
        return ""
    if isinstance(dt_value, str):
        try:
            dt = datetime.fromisoformat(dt_value.replace('Z', '+00:00'))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(IST).strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            return dt_value
    if isinstance(dt_value, datetime):
        if dt_value.tzinfo is None:
            dt_value = dt_value.replace(tzinfo=timezone.utc)
        return dt_value.astimezone(IST).strftime('%Y-%m-%d %H:%M:%S')
    return str(dt_value)


async def export_conversation_to_excel():
    """Export a single conversation to Excel."""
    try:
        if not CONVERSATION_ID:
            raise ValueError("CONVERSATION_ID is not set. Please set it at the top of the script.")

        conv_id = str(CONVERSATION_ID)
        logger.info(f"Exporting conversation ID: {conv_id}")

        # Fetch all data
        logger.info("Fetching all data from backend...")
        res = await backend_client.fetch_all_data()
        data = res.get("data", res)

        conversations = data.get('conversations', [])
        messages      = data.get('messages', [])
        users         = data.get('users', [])

        logger.info(
            f"Retrieved {len(conversations)} conversations, "
            f"{len(messages)} messages, {len(users)} users"
        )

        # Find the target conversation
        conversation = next(
            (c for c in conversations if str(c.get('id')) == conv_id),
            None
        )

        if not conversation:
            raise ValueError(f"Conversation with ID '{conv_id}' not found in the data.")

        logger.info(f"Found conversation: {conversation}")

        # Build user lookup
        user_lookup = {user.get('id'): user for user in users}

        # Build workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Conversations"

        headers = [
            "Coach ID",
            "User ID",
            "Conversation ID",
            "User Converted",
            "Messages",
            "Started At",
            "Last Message At",
        ]

        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Build the single row
        user_id         = conversation.get('userId')
        coach_id        = conversation.get('coachId')
        chatroom_id     = conversation.get('chatroomId')
        started_at      = conversation.get('startedAt')
        last_message_at = conversation.get('lastMessageAt')

        user_data       = user_lookup.get(user_id)
        user_converted  = get_user_converted_status(user_data)
        formatted_msgs  = format_messages(messages, conv_id)

        ws.append([
            coach_id or "",
            user_id or "",
            chatroom_id or "",
            user_converted,
            formatted_msgs,
            format_datetime(started_at),
            format_datetime(last_message_at),
        ])

        # Column widths
        ws.column_dimensions['A'].width = 12   # Coach ID
        ws.column_dimensions['B'].width = 20   # User ID
        ws.column_dimensions['C'].width = 25   # Conversation ID
        ws.column_dimensions['D'].width = 15   # User Converted
        ws.column_dimensions['E'].width = 80   # Messages
        ws.column_dimensions['F'].width = 20   # Started At
        ws.column_dimensions['G'].width = 20   # Last Message At

        # Wrap text for Messages column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        filename = f"conversation_{conv_id}_{timestamp}.xlsx"

        wb.save(filename)
        logger.info(f"Successfully exported conversation to {filename}")

        return filename

    except Exception as e:
        logger.error(f"Error exporting conversation: {e}", exc_info=True)
        raise
    finally:
        await backend_client.close()


async def main():
    """Entry point for the script."""
    try:
        filename = await export_conversation_to_excel()
        print(f"\n{'='*60}")
        print(f"Export completed successfully!")
        print(f"Conversation ID : {CONVERSATION_ID}")
        print(f"File            : {filename}")
        print(f"{'='*60}\n")
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"Export failed: {e}")
        print(f"{'='*60}\n")
        exit(1)


if __name__ == "__main__":
    asyncio.run(main())
