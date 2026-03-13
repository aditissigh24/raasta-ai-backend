"""
Script to export batch-specific conversation data to Excel.

Filters conversations by a selected batch (morning / evening / night) and a
date range.  Each batch automatically applies its IST time window:
    morning : 05:00 – 13:00 IST
    evening : 13:00 – 21:00 IST
    night   : 21:00 – 05:00 IST (spans midnight)

Each batch has a fixed list of human coaches assigned to it.
Coach type (AI_COACH / HUMAN_COACH) is derived from the senderType field of
the conversation's messages.

Usage:
    python batch_conversation_excel.py

Configuration (edit the variables below):
    BATCH      - "morning", "evening", or "night"
    START_DATE - inclusive start date "YYYY-MM-DD" in IST, or None
    END_DATE   - inclusive end date "YYYY-MM-DD" in IST, or None

Requirements:
    - Backend API must be running and accessible
    - BACKEND_BASE_URL and BACKEND_API_KEY must be configured in .env file
    - openpyxl package must be installed (pip install openpyxl)

Output:
    Creates an Excel file named: batch_conversations_{BATCH}_{YYYY-MM-DD_HHMMSS}.xlsx

    Columns:
    1.  Coach ID        - Database ID of the coach
    2.  Coach Name      - Full name of the coach (firstName + lastName)
    3.  Coach Type      - AI_COACH or HUMAN_COACH (from message senderType)
    4.  User ID         - CometChat UID of the user
    5.  Conversation ID - CometChat chatroom ID
    6.  User Converted  - "Yes" if user converted (isGuest=false), "No" if guest
    7.  Messages        - Formatted conversation messages (coach: text\\nuser: text)
    8.  Started At      - When the conversation started
    9.  Last Message At - When the last message was sent

API Endpoint:
    GET /api/admin/all-data
    - Returns: {conversations: [], messages: [], users: [], coaches: []}
    - Requires: x-api-key header with BACKEND_API_KEY
"""
import asyncio
import logging
from datetime import datetime, timezone, timedelta, time, date
from typing import Dict, List, Optional, Set, Tuple

# IST = UTC + 5:30
IST = timezone(timedelta(hours=5, minutes=30))
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from services.backend_client import backend_client

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================
# BATCH CONFIGURATION
# ============================
# Set BATCH to one of: "morning", "evening", "night"
# Each batch automatically applies its IST time window:
#   morning : 05:00 – 13:00 IST
#   evening : 13:00 – 21:00 IST
#   night   : 21:00 – 05:00 IST (spans midnight)
BATCH = "night"

# Time windows per batch in IST (start_time inclusive, end_time exclusive)
# night wraps midnight: 21:00 of day D → 05:00 of day D+1
BATCH_TIME_RANGES: Dict[str, Tuple[time, time]] = {
    "morning": (time(5, 0),  time(13, 0)),
    "evening": (time(13, 0), time(21, 0)),
    "night":   (time(21, 0), time(5, 0)),   # wraps past midnight
}

# Coaches assigned to each batch (matched by firstName + " " + lastName)
BATCH_COACHES: Dict[str, List[str]] = {
    "morning": [
        "Aarushi Khanna",
        "Dhriti Singh",
        "Pooja Arora",
        "Ananya Sharma",
        "Sanya Malhotra",
        "Palak Arora"
    ],
    "evening": [
        "Shivani Agarwal",
        "Ishita Verma",
        "Tanya Bansal",
        "Kavya Gupta",
        "Neha Kapoor"
    ],
    "night": [
        "Meera Iyer",
        "Rhea Mehta"
    ],
}
# ============================

# ============================
# DATE RANGE FILTER CONFIGURATION
# ============================
# Enter dates only (no time) — the batch's time window is applied automatically.
# Format: "YYYY-MM-DD" or None to disable filtering
START_DATE = "2026-03-09"   # inclusive
END_DATE   = "2026-03-10"   # inclusive
# ============================


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def format_messages(messages: List[Dict], conversation_id: str) -> str:
    """
    Format messages for a conversation into a readable string.

    Args:
        messages: List of all message dictionaries
        conversation_id: The conversation ID to filter messages for

    Returns:
        Formatted string with messages like "coach: text\\nuser: text"
    """
    conv_messages = [
        m for m in messages if str(m.get('conversationId')) == str(conversation_id)
    ]
    conv_messages.sort(key=lambda x: x.get('createdAt', x.get('id', '')))

    formatted_lines = []
    for msg in conv_messages:
        sender_type = msg.get('senderType', '').upper()
        text = msg.get('text', '')

        if not text:
            continue

        prefix = 'coach' if sender_type in ('AI_COACH', 'HUMAN_COACH') else 'user'
        formatted_lines.append(f"{prefix}: {text}")

    return '\n'.join(formatted_lines)


def get_coach_type(messages: List[Dict], conversation_id: str) -> str:
    """
    Derive the coach type for a conversation from its messages.

    Scans messages for the first one where senderType is AI_COACH or
    HUMAN_COACH and returns that value.

    Args:
        messages: List of all message dictionaries
        conversation_id: The conversation ID to scan

    Returns:
        "AI_COACH", "HUMAN_COACH", or "UNKNOWN"
    """
    conv_messages = [
        m for m in messages if str(m.get('conversationId')) == str(conversation_id)
    ]
    conv_messages.sort(key=lambda x: x.get('createdAt', x.get('id', '')))

    for msg in conv_messages:
        sender_type = msg.get('senderType', '').upper()
        if sender_type in ('AI_COACH', 'HUMAN_COACH'):
            return sender_type

    return "UNKNOWN"


def get_user_converted_status(user_data: Optional[Dict]) -> str:
    """
    Determine user conversion status.

    Args:
        user_data: User dictionary with isGuest field

    Returns:
        "Yes" if user converted (isGuest=false), "No" otherwise
    """
    if not user_data:
        return "Unknown"

    is_guest = user_data.get('isGuest')

    if is_guest is True:
        return "No"
    elif is_guest is False:
        return "Yes"
    return "Unknown"


def format_datetime(dt_value) -> str:
    """
    Format a datetime value for Excel display, converted to IST (UTC+5:30).

    Args:
        dt_value: DateTime string or value (stored as UTC in the database)

    Returns:
        Formatted datetime string in IST
    """
    if not dt_value:
        return ""

    if isinstance(dt_value, str):
        try:
            dt = datetime.fromisoformat(dt_value.replace('Z', '+00:00'))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            dt_ist = dt.astimezone(IST)
            return dt_ist.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            return dt_value

    if isinstance(dt_value, datetime):
        if dt_value.tzinfo is None:
            dt_value = dt_value.replace(tzinfo=timezone.utc)
        return dt_value.astimezone(IST).strftime('%Y-%m-%d %H:%M:%S')

    return str(dt_value)


def parse_datetime(dt_value) -> Optional[datetime]:
    """
    Parse a datetime value to a timezone-aware datetime object.
    Naive strings (no tz info) are assumed to be UTC (used for DB values).

    Args:
        dt_value: DateTime string or value

    Returns:
        Timezone-aware datetime object, or None if parsing fails
    """
    if not dt_value:
        return None

    if isinstance(dt_value, datetime):
        if dt_value.tzinfo is None:
            return dt_value.replace(tzinfo=timezone.utc)
        return dt_value

    if isinstance(dt_value, str):
        try:
            return datetime.fromisoformat(dt_value.replace('Z', '+00:00'))
        except Exception:
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y-%m-%dT%H:%M:%S']:
                try:
                    dt = datetime.strptime(dt_value, fmt)
                    return dt.replace(tzinfo=timezone.utc)
                except Exception:
                    continue

    return None


def parse_date(date_str: Optional[str]) -> Optional[date]:
    """
    Parse a date string "YYYY-MM-DD" to a date object.

    Args:
        date_str: Date string e.g. "2026-02-17"

    Returns:
        date object or None if parsing fails
    """
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
    except Exception:
        return None


def is_in_batch_window(
    conversation: Dict,
    batch: str,
    start_date: Optional[str],
    end_date: Optional[str],
) -> bool:
    """
    Check if a conversation's startedAt (in IST) falls within the given date
    range AND within the batch's IST time window.

    For morning / evening the time window is within a single calendar day.
    For night (21:00–05:00) the window spans midnight: the IST date checked
    against START_DATE/END_DATE is the date when the shift *starts* (21:00 side).

    Args:
        conversation: Conversation dictionary
        batch: "morning", "evening", or "night"
        start_date: Inclusive start date string "YYYY-MM-DD" in IST, or None
        end_date: Inclusive end date string "YYYY-MM-DD" in IST, or None

    Returns:
        True if the conversation falls within the batch window and date range
    """
    conv_dt_utc = parse_datetime(conversation.get('startedAt'))
    if not conv_dt_utc:
        return False

    if conv_dt_utc.tzinfo is None:
        conv_dt_utc = conv_dt_utc.replace(tzinfo=timezone.utc)

    # Convert to IST for all comparisons
    conv_dt_ist = conv_dt_utc.astimezone(IST)
    conv_time   = conv_dt_ist.time()
    conv_date   = conv_dt_ist.date()

    # --- Batch time window check ---
    batch_start, batch_end = BATCH_TIME_RANGES[batch]

    if batch == "night":
        # Night wraps midnight: valid if time >= 21:00 OR time < 05:00
        in_window = conv_time >= batch_start or conv_time < batch_end
        # For date-range purposes, attribute night conversations to the
        # shift-start date: if it's after midnight (< 05:00), the shift
        # started the previous calendar day
        if conv_time < batch_end:
            shift_date = conv_date - timedelta(days=1)
        else:
            shift_date = conv_date
    else:
        in_window  = batch_start <= conv_time < batch_end
        shift_date = conv_date

    if not in_window:
        return False

    # --- Date range check (against IST shift date) ---
    start = parse_date(start_date)
    end   = parse_date(end_date)

    if start and shift_date < start:
        return False
    if end and shift_date > end:
        return False

    return True


def resolve_batch_coach_ids(
    coaches: List[Dict],
    batch_coach_names: List[str],
) -> Set[int]:
    """
    Return the set of coach DB IDs whose name matches an entry in batch_coach_names.

    Matching rules (case-insensitive, whitespace-stripped):
    - If the entry contains a space it is treated as a full name
      (firstName + " " + lastName) and matched exactly.
    - If the entry is a single word it is matched against firstName only,
      allowing "Shivani" to match "Shivani Agarwal".

    Args:
        coaches: List of coach dictionaries from the API
        batch_coach_names: List of names (full or first-name-only) for the batch

    Returns:
        Set of integer coach IDs
    """
    full_names  = {n.strip().lower() for n in batch_coach_names if ' ' in n.strip()}
    first_names = {n.strip().lower() for n in batch_coach_names if ' ' not in n.strip()}

    coach_ids: Set[int] = set()

    for coach in coaches:
        first = (coach.get('firstName') or '').strip()
        last  = (coach.get('lastName') or '').strip()
        full_name = f"{first} {last}".strip().lower()

        if full_name in full_names or first.lower() in first_names:
            coach_id = coach.get('id')
            if coach_id is not None:
                coach_ids.add(coach_id)

    return coach_ids


def build_coach_name_lookup(coaches: List[Dict]) -> Dict[int, str]:
    """
    Build a mapping from coach DB ID to full display name.

    Args:
        coaches: List of coach dictionaries from the API

    Returns:
        Dict mapping coach ID (int) -> "FirstName LastName"
    """
    lookup: Dict[int, str] = {}
    for coach in coaches:
        coach_id = coach.get('id')
        if coach_id is None:
            continue
        first = coach.get('firstName', '') or ''
        last = coach.get('lastName', '') or ''
        lookup[coach_id] = f"{first} {last}".strip()
    return lookup


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

async def export_batch_conversations_to_excel():
    """Export conversations for the selected batch to Excel."""
    try:
        batch = BATCH.strip().lower()

        if batch not in BATCH_COACHES:
            raise ValueError(
                f"Invalid BATCH value '{BATCH}'. "
                f"Must be one of: {', '.join(BATCH_COACHES.keys())}"
            )

        batch_coach_names = BATCH_COACHES[batch]
        batch_start_t, batch_end_t = BATCH_TIME_RANGES[batch]
        logger.info(
            f"Batch: {batch} | Time window (IST): "
            f"{batch_start_t.strftime('%H:%M')} – {batch_end_t.strftime('%H:%M')} | "
            f"Coaches: {batch_coach_names}"
        )

        if START_DATE or END_DATE:
            logger.info(
                f"Date range filter (IST): {START_DATE or 'No start'} "
                f"to {END_DATE or 'No end'}"
            )
        else:
            logger.info("No date range filter applied — exporting all conversations")

        # Fetch all data
        logger.info("Fetching all data from backend...")
        res = await backend_client.fetch_all_data()
        data = res.get("data", res)  # handle both wrapped and unwrapped responses

        conversations = data.get('conversations', [])
        messages     = data.get('messages', [])
        users        = data.get('users', [])
        ratings      = data.get('ratings', [])

        # /api/admin/all-data does not include coaches; fetch them separately
        logger.info("Fetching coaches from backend...")
        coaches = await backend_client.fetch_coaches()

        logger.info(
            f"Retrieved {len(conversations)} conversations, "
            f"{len(messages)} messages, {len(users)} users, "
            f"{len(ratings)} ratings, {len(coaches)} coaches"
        )

        # Log all fetched coach names to help debug mismatches
        fetched_names = [
            f"{c.get('firstName', '')} {c.get('lastName', '')}".strip()
            for c in coaches
        ]
        logger.info(f"All coach names from API: {fetched_names}")

        # Resolve coach IDs for this batch
        batch_coach_ids = resolve_batch_coach_ids(coaches, batch_coach_names)
        logger.info(
            f"Resolved {len(batch_coach_ids)} coach DB IDs for batch '{batch}': "
            f"{batch_coach_ids}"
        )

        if not batch_coach_ids:
            logger.warning(
                "No coach IDs resolved for this batch. "
                "Check that coach names match firstName + lastName in the database. "
                f"API names: {fetched_names}"
            )

        # Build coach name lookup
        coach_name_lookup = build_coach_name_lookup(coaches)

        # Build user lookup
        user_lookup = {user.get('id'): user for user in users}

        # Build rating lookup: (userId, coachId) -> highest rating value
        # Keyed by (userId, coachId) since conversationId/sessionId on ratings
        # are unreliable. If multiple ratings exist for the same user+coach pair,
        # keep the highest value.
        rating_lookup: Dict[tuple, int] = {}
        for r in ratings:
            uid = r.get('userId')
            cid = r.get('coachId')
            val = r.get('rating')
            if uid is None or cid is None or val is None:
                continue
            key = (str(uid), int(cid))
            if key not in rating_lookup or val > rating_lookup[key]:
                rating_lookup[key] = val

        # Filter conversations: batch time window + date range + batch coaches
        filtered_conversations = [
            conv for conv in conversations
            if is_in_batch_window(conv, batch, START_DATE, END_DATE)
            and conv.get('coachId') in batch_coach_ids
        ]

        logger.info(
            f"Filtered to {len(filtered_conversations)} conversations "
            f"for batch '{batch}' within date range"
        )

        # Build Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{batch.capitalize()} Batch"

        headers = [
            "Coach ID",
            "Coach Name",
            "Coach Type",
            "User ID",
            "Conversation ID",
            "User Converted",
            "Rating",
            "Messages",
            "Started At",
            "Last Message At",
        ]

        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        logger.info("Processing conversations...")
        row_count = 0

        for conv in filtered_conversations:
            conv_id        = conv.get('id')
            user_id        = conv.get('userId')
            coach_id       = conv.get('coachId')
            chatroom_id    = conv.get('chatroomId')
            started_at     = conv.get('startedAt')
            last_message_at = conv.get('lastMessageAt')

            coach_name      = coach_name_lookup.get(coach_id, "") if coach_id else ""
            coach_type      = get_coach_type(messages, conv_id)
            if coach_type != 'HUMAN_COACH':
                continue
            user_data       = user_lookup.get(user_id)
            user_converted  = get_user_converted_status(user_data)
            rating          = rating_lookup.get((str(user_id), int(coach_id)), None) if user_id and coach_id else None
            formatted_msgs  = format_messages(messages, conv_id)

            ws.append([
                coach_id or "",
                coach_name,
                coach_type,
                user_id or "",
                chatroom_id or "",
                user_converted,
                rating,          # None becomes an empty cell in Excel (treated as null)
                formatted_msgs,
                format_datetime(started_at),
                format_datetime(last_message_at),
            ])

            row_count += 1
            if row_count % 10 == 0:
                logger.info(f"Processed {row_count} conversations...")

        # Column widths
        ws.column_dimensions['A'].width = 12   # Coach ID
        ws.column_dimensions['B'].width = 22   # Coach Name
        ws.column_dimensions['C'].width = 15   # Coach Type
        ws.column_dimensions['D'].width = 20   # User ID
        ws.column_dimensions['E'].width = 25   # Conversation ID
        ws.column_dimensions['F'].width = 15   # User Converted
        ws.column_dimensions['G'].width = 10   # Rating
        ws.column_dimensions['H'].width = 80   # Messages
        ws.column_dimensions['I'].width = 20   # Started At
        ws.column_dimensions['J'].width = 20   # Last Message At

        # Wrap text in Messages column (H)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        filename = f"batch_conversations_{batch}_{timestamp}.xlsx"

        wb.save(filename)
        logger.info(f"Successfully exported {row_count} conversations to {filename}")

        return filename

    except Exception as e:
        logger.error(f"Error exporting conversations: {e}", exc_info=True)
        raise
    finally:
        await backend_client.close()


async def main():
    """Entry point for the script."""
    try:
        filename = await export_batch_conversations_to_excel()
        print(f"\n{'='*60}")
        print(f"Export completed successfully!")
        print(f"Batch : {BATCH}")
        print(f"File  : {filename}")
        print(f"{'='*60}\n")
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"Export failed: {e}")
        print(f"{'='*60}\n")
        exit(1)


if __name__ == "__main__":
    asyncio.run(main())
