"""
Script to export conversation data to Excel.

This script fetches all conversations, messages, and user data from the backend
and exports them to an Excel file with formatted messages.

Usage:
    python conversation_excel.py

Requirements:
    - Backend API must be running and accessible
    - BACKEND_BASE_URL and BACKEND_API_KEY must be configured in .env file
    - openpyxl package must be installed (pip install openpyxl)

Output:
    Creates an Excel file named: conversations_YYYY-MM-DD_HHMMSS.xlsx
    
    Columns:
    1. Coach ID - Database ID of the coach
    2. User ID - CometChat UID of the user
    3. Conversation ID - CometChat chatroom ID
    4. User Converted - "Yes" if user converted (isGuest=false), "No" if guest (isGuest=true)
    5. Messages - Formatted conversation messages (coach: text\\nuser: text)
    6. Started At - When the conversation started
    7. Last Message At - When the last message was sent

API Endpoint:
    GET /api/admin/all-data
    - Returns: {conversations: [], messages: [], users: [], coaches: []}
    - Requires: x-api-key header with BACKEND_API_KEY
"""
import asyncio
import logging
from datetime import datetime, timezone
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from services.backend_client import backend_client

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================
# DATE RANGE FILTER CONFIGURATION
# ============================
# Set these variables to filter conversations by datetime range
# Format: "YYYY-MM-DD HH:MM:SS" or None to disable filtering
# Example: "2026-02-01 00:00:00"
START_DATETIME = "2026-02-16 00:00:00"  # Start of the date range (inclusive)
END_DATETIME = "2026-02-16 19:00:00"    # End of the date range (inclusive)
# ============================


def format_messages(messages: List[Dict], conversation_id: str) -> str:
    """
    Format messages for a conversation into a readable string.
    
    Args:
        messages: List of message dictionaries
        conversation_id: The conversation ID to filter messages
        
    Returns:
        Formatted string with messages like "coach: text\\nuser: text"
    """
    # Filter messages for this conversation
    conv_messages = [m for m in messages if str(m.get('conversationId')) == str(conversation_id)]
    
    # Sort by createdAt or id to ensure chronological order
    conv_messages.sort(key=lambda x: x.get('createdAt', x.get('id', '')))
    
    formatted_lines = []
    for msg in conv_messages:
        sender_type = msg.get('senderType', '').upper()
        text = msg.get('text', '')
        
        if not text:
            continue
        
        # Determine if it's a coach or user message from senderType
        # Coach messages can be either AI_COACH or HUMAN_COACH
        if sender_type in ('AI_COACH', 'HUMAN_COACH'):
            prefix = 'coach'
        else:
            prefix = 'user'
        
        formatted_lines.append(f"{prefix}: {text}")
    
    return '\n'.join(formatted_lines)


def get_user_converted_status(user_data: Optional[Dict]) -> str:
    """
    Determine user conversion status from user data.
    
    Args:
        user_data: User dictionary with isGuest field
        
    Returns:
        "Yes" if user converted (isGuest=false), "No" otherwise
    """
    if not user_data:
        return "Unknown"
    
    is_guest = user_data.get('isGuest')
    
    # If isGuest is True, user didn't convert -> "No"
    # If isGuest is False, user converted -> "Yes"
    if is_guest is True:
        return "No"
    elif is_guest is False:
        return "Yes"
    else:
        return "Unknown"


def format_datetime(dt_value) -> str:
    """
    Format datetime value for Excel.
    
    Args:
        dt_value: DateTime string or value
        
    Returns:
        Formatted datetime string
    """
    if not dt_value:
        return ""
    
    # If it's already a string, return it
    if isinstance(dt_value, str):
        try:
            # Try to parse and reformat for consistency
            dt = datetime.fromisoformat(dt_value.replace('Z', '+00:00'))
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except:
            return dt_value
    
    return str(dt_value)


def parse_datetime(dt_value) -> Optional[datetime]:
    """
    Parse datetime value to datetime object (timezone-aware).
    
    Args:
        dt_value: DateTime string or value
        
    Returns:
        Timezone-aware datetime object or None if parsing fails
    """
    if not dt_value:
        return None
    
    if isinstance(dt_value, datetime):
        # If already a datetime, ensure it's timezone-aware
        if dt_value.tzinfo is None:
            # Make it UTC if naive
            return dt_value.replace(tzinfo=timezone.utc)
        return dt_value
    
    if isinstance(dt_value, str):
        try:
            # Try ISO format with Z (timezone-aware)
            return datetime.fromisoformat(dt_value.replace('Z', '+00:00'))
        except:
            try:
                # Try common datetime formats (will be naive, so add UTC)
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y-%m-%dT%H:%M:%S']:
                    try:
                        dt = datetime.strptime(dt_value, fmt)
                        # Make timezone-aware (assume UTC)
                        return dt.replace(tzinfo=timezone.utc)
                    except:
                        continue
            except:
                pass
    
    return None


def is_in_date_range(conversation: Dict, start_dt: Optional[str], end_dt: Optional[str]) -> bool:
    """
    Check if a conversation falls within the specified date range.
    
    Args:
        conversation: Conversation dictionary
        start_dt: Start datetime string (inclusive) or None
        end_dt: End datetime string (inclusive) or None
        
    Returns:
        True if conversation is within range (or no range specified), False otherwise
    """
    # If no date range specified, include all conversations
    if not start_dt and not end_dt:
        return True
    
    # Parse conversation startedAt
    conv_started_at = conversation.get('startedAt')
    conv_dt = parse_datetime(conv_started_at)
    
    if not conv_dt:
        # If conversation has no date, exclude it when filtering
        return False if (start_dt or end_dt) else True
    
    # Ensure conv_dt is timezone-aware (in case parse_datetime didn't add timezone)
    if conv_dt.tzinfo is None:
        conv_dt = conv_dt.replace(tzinfo=timezone.utc)
    
    # Parse start datetime
    if start_dt:
        start_datetime = parse_datetime(start_dt)
        if start_datetime:
            # Ensure start_datetime is timezone-aware
            if start_datetime.tzinfo is None:
                start_datetime = start_datetime.replace(tzinfo=timezone.utc)
            if conv_dt < start_datetime:
                return False
    
    # Parse end datetime
    if end_dt:
        end_datetime = parse_datetime(end_dt)
        if end_datetime:
            # Ensure end_datetime is timezone-aware
            if end_datetime.tzinfo is None:
                end_datetime = end_datetime.replace(tzinfo=timezone.utc)
            if conv_dt > end_datetime:
                return False
    
    return True


async def export_conversations_to_excel():
    """
    Main function to export conversation data to Excel.
    """
    try:
        logger.info("Starting conversation export...")
        
        # Log date range filter if configured
        if START_DATETIME or END_DATETIME:
            logger.info(f"Date range filter: {START_DATETIME or 'No start'} to {END_DATETIME or 'No end'}")
        else:
            logger.info("No date range filter applied - exporting all conversations")
        
        # Fetch all data from backend
        logger.info("Fetching all data from backend...")
        res = await backend_client.fetch_all_data()
        data = res.get("data")
        
        conversations = data.get('conversations', [])
        messages = data.get('messages', [])
        users = data.get('users', [])
        
        logger.info(f"Retrieved {len(conversations)} conversations, "
                   f"{len(messages)} messages, {len(users)} users")
        
        # Filter conversations by date range
        filtered_conversations = [
            conv for conv in conversations 
            if is_in_date_range(conv, START_DATETIME, END_DATETIME)
        ]
        
        if START_DATETIME or END_DATETIME:
            logger.info(f"Filtered to {len(filtered_conversations)} conversations within date range")
        
        # Create user lookup dictionary for isGuest status
        user_lookup = {user.get('id'): user for user in users}
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Conversations"
        
        # Define headers
        headers = [
            "Coach ID",
            "User ID",
            "Conversation ID",
            "User Converted",
            "Messages",
            "Started At",
            "Last Message At"
        ]
        
        # Write headers with formatting
        ws.append(headers)
        header_row = ws[1]
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Process each conversation
        logger.info("Processing conversations...")
        row_count = 0
        
        for conv in filtered_conversations:
            conv_id = conv.get('id')
            user_id = conv.get('userId')
            coach_id = conv.get('coachId')
            chatroom_id = conv.get('chatroomId')
            started_at = conv.get('startedAt')
            last_message_at = conv.get('lastMessageAt')
            
            # Get user data from the simplified users array
            user_data = user_lookup.get(user_id)
            user_converted = get_user_converted_status(user_data)
            
            # Format messages
            formatted_messages = format_messages(messages, conv_id)
            
            # Create row
            row = [
                coach_id or "",
                user_id or "",
                chatroom_id or "",
                user_converted,
                formatted_messages,
                format_datetime(started_at),
                format_datetime(last_message_at)
            ]
            
            ws.append(row)
            row_count += 1
            
            if row_count % 10 == 0:
                logger.info(f"Processed {row_count} conversations...")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12  # Coach ID
        ws.column_dimensions['B'].width = 20  # User ID
        ws.column_dimensions['C'].width = 25  # Conversation ID
        ws.column_dimensions['D'].width = 15  # User Converted
        ws.column_dimensions['E'].width = 80  # Messages
        ws.column_dimensions['F'].width = 20  # Started At
        ws.column_dimensions['G'].width = 20  # Last Message At
        
        # Set text wrapping for messages column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        filename = f"conversations_{timestamp}.xlsx"
        
        # Save workbook
        wb.save(filename)
        logger.info(f"✅ Successfully exported {row_count} conversations to {filename}")
        
        return filename
        
    except Exception as e:
        logger.error(f"❌ Error exporting conversations: {e}", exc_info=True)
        raise
    finally:
        await backend_client.close()


async def main():
    """Entry point for the script."""
    try:
        filename = await export_conversations_to_excel()
        print(f"\n{'='*60}")
        print(f"Export completed successfully!")
        print(f"File: {filename}")
        print(f"{'='*60}\n")
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"Export failed: {e}")
        print(f"{'='*60}\n")
        exit(1)


if __name__ == "__main__":
    asyncio.run(main())

