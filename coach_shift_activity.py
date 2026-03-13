"""
Script to export coach shift activity to Excel.

For each day from Feb 17 to Mar 10 (IST), queries the MongoDB `coach-events`
collection for `coach_auth_session_started` events and shows which coaches
were active in each shift:
  - Morning : 05:00–13:00 IST
  - Evening : 13:00–21:00 IST
  - Night   : 21:00–05:00 IST (next calendar day)

Each cell lists real coach card names (one per line).

Fill-in rule:
  - If a coach was active in their assigned shift on a given day, they are
    shown only in their assigned shift column — even if they also have events
    in other shifts that day.
  - If a coach has NO event in their assigned shift on a given day but does
    have events in another shift, they appear in that other shift's column
    marked "(fill-in)".

Usage:
    python coach_shift_activity.py

Output:
    coach_shift_activity_YYYY-MM-DD_HHMMSS.xlsx
"""
import asyncio
import logging
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from typing import Dict, List, Set, Tuple

from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.settings import settings

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

IST_OFFSET = timedelta(hours=5, minutes=30)

# Date range (inclusive, IST calendar dates)
RANGE_START = date(2026, 2, 17)
RANGE_END = date(2026, 3, 10)

SHIFT_MORNING = "Morning"
SHIFT_EVENING = "Evening"
SHIFT_NIGHT = "Night"

SHIFTS = [SHIFT_MORNING, SHIFT_EVENING, SHIFT_NIGHT]

# ---------------------------------------------------------------------------
# Coach mapping: email (lower-case) -> (assigned_shift, coach_card_name_in_db, real_name)
# coach_card_name_in_db is the fake name stored in event_properties.coach_name
# real_name is the actual coach name shown in the Excel
# ---------------------------------------------------------------------------

# Tuple: (assigned_shift, coach_card_name_in_db, real_name)
COACH_MAP: Dict[str, Tuple[str, str, str]] = {
    # Morning coaches
    "ramandeep@sheroesmars.com":   (SHIFT_MORNING, "Aarushi Khanna",   "Ramandeep"),
    "farzana@sheroesmars.com":     (SHIFT_MORNING, "Dhriti Singh",     "Farzana"),
    "konika1@sheroesmars.com":     (SHIFT_MORNING, "Pooja Arora",      "Konika"),
    "puspa@sheroesmars.com":       (SHIFT_MORNING, "Ananya Sharma",    "Puspa"),
    "fatima@sheroesmars.com":      (SHIFT_MORNING, "Sanya Malhotra",   "Fatima"),
    "soham@sheroesmars.com":       (SHIFT_MORNING, "Palak Arora",      "Soham"),
    # Evening coaches
    "shefali@sheroesmars.com":             (SHIFT_EVENING, "Shivani Agarwal",  "Shefali"),
    "noureen@sheroesmars.com":             (SHIFT_EVENING, "Ishita Verma",     "Noureen"),
    "mitali@sheroesmars.com":              (SHIFT_EVENING, "Tanya Bansal",     "Mitali"),
    "komalpreet@sheroesmars.com":          (SHIFT_EVENING, "Kavya Gupta",      "Komalpreet"),
    "prashanthi@sheroesmars.com":          (SHIFT_EVENING, "Neha Kapoor",      "Prashanthi"),
    "farida.khatoon@sheroesmars.com":      (SHIFT_EVENING, "Sonal Chawla",     "Farida"),
    "shuchismita@sheroesmars.com":         (SHIFT_EVENING, "Aditi Mukherjee",  "Shuchismita"),
    "manasa@sheroesmars.com":              (SHIFT_EVENING, "Lavanya Joshi",    "Manasa"),
    # Night coaches
    "sarabjeet@sheroesmars.com":   (SHIFT_NIGHT, "Meera Iyer",  "Sarabjeet"),
    "adiba@sheroesmars.com":       (SHIFT_NIGHT, "Rhea Mehta",  "Adiba"),
}


def _to_ist(utc_dt: datetime) -> datetime:
    """Convert a UTC datetime to IST (naive or timezone-aware both handled)."""
    if utc_dt.tzinfo is None:
        utc_dt = utc_dt.replace(tzinfo=timezone.utc)
    return utc_dt.astimezone(timezone(IST_OFFSET))


def _shift_for_ist_time(ist_dt: datetime) -> Tuple[str, date]:
    """
    Given an IST datetime, return (shift_name, calendar_date_of_shift).

    Night shift spans 21:00 on day D to 05:00 on day D+1.
    The "shift date" for Night is the calendar day when the shift *starts*
    (i.e. 21:00 side), so an event at 01:00 on Feb 18 belongs to the Night
    shift of Feb 17.
    """
    h = ist_dt.hour
    d = ist_dt.date()

    if 5 <= h < 13:
        return SHIFT_MORNING, d
    elif 13 <= h < 21:
        return SHIFT_EVENING, d
    else:
        # Night: 21:00–05:00
        # If hour >= 21, shift date = today
        # If hour < 5, shift date = yesterday
        if h >= 21:
            return SHIFT_NIGHT, d
        else:
            return SHIFT_NIGHT, d - timedelta(days=1)


async def fetch_events() -> List[Dict]:
    """Query MongoDB coach-events for both session_started and session_ended events."""
    client = AsyncIOMotorClient(settings.MONGODB_URI)
    try:
        db = client[settings.MONGODB_DATABASE]
        collection = db["coach_events"]

        # Convert IST range boundaries to UTC for the query
        start_utc = datetime(
            RANGE_START.year, RANGE_START.month, RANGE_START.day,
            0, 0, 0, tzinfo=timezone(IST_OFFSET)
        ).astimezone(timezone.utc).replace(tzinfo=None)

        # End: Mar 10 23:59:59 IST → include full night shift ending 05:00 Mar 11
        # We extend to Mar 11 05:00 IST to capture the tail of the last Night shift
        end_ist = datetime(
            RANGE_END.year, RANGE_END.month, RANGE_END.day + 1,
            5, 0, 0, tzinfo=timezone(IST_OFFSET)
        )
        end_utc = end_ist.astimezone(timezone.utc).replace(tzinfo=None)

        query = {
            "event_name": {"$in": ["coach_auth_session_started", "coach_auth_session_ended"]},
            "created_at": {"$gte": start_utc, "$lte": end_utc},
        }

        logger.info(f"Querying coach-events from {start_utc} to {end_utc} UTC")
        cursor = collection.find(query, {"event_name": 1, "event_properties": 1, "created_at": 1})
        events = await cursor.to_list(length=None)
        started = sum(1 for e in events if e.get("event_name") == "coach_auth_session_started")
        ended = sum(1 for e in events if e.get("event_name") == "coach_auth_session_ended")
        logger.info(f"Fetched {len(events)} events ({started} started, {ended} ended)")
        return events
    finally:
        client.close()
def compute_most_active_shift(
    events: List[Dict],
    fake_to_real: Dict[str, str],
) -> Dict[str, Dict[date, str]]:
    """
    From coach_auth_session_ended events, sum session_duration seconds per
    (real_name, shift_date, shift_name) and return the shift with the highest
    total seconds for each coach per day.

    Returns: { real_name: { shift_date: shift_name_with_max_seconds } }
    Only coaches/days with at least one ended event appear in the result.
    """
    # duration_sums[real_name][shift_date][shift_name] = total seconds
    duration_sums: Dict[str, Dict[date, Dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(float))
    )

    for event in events:
        if event.get("event_name") != "coach_auth_session_ended":
            continue

        props = event.get("event_properties") or {}
        coach_name_raw = (props.get("coach_name") or "").strip()
        if not coach_name_raw or coach_name_raw.lower() == "test coach":
            continue

        session_duration = props.get("session_duration")
        if session_duration is None:
            continue
        try:
            seconds = float(session_duration)
        except (TypeError, ValueError):
            continue
        if seconds <= 0:
            continue

        created_at = event.get("created_at")
        if not created_at or not isinstance(created_at, datetime):
            continue

        ist_dt = _to_ist(created_at)
        shift_name, shift_date = _shift_for_ist_time(ist_dt)

        if shift_date < RANGE_START or shift_date > RANGE_END:
            continue

        real_name = fake_to_real.get(coach_name_raw.lower(), coach_name_raw)
        duration_sums[real_name][shift_date][shift_name] += seconds

    # Collapse to the shift with max seconds per (coach, day)
    result: Dict[str, Dict[date, str]] = {}
    for real_name, day_map in duration_sums.items():
        result[real_name] = {}
        for shift_date, shift_seconds in day_map.items():
            best_shift = max(shift_seconds, key=shift_seconds.__getitem__)
            result[real_name][shift_date] = best_shift

    return result


def bucket_events(
    events: List[Dict],
    most_active_shift: Dict[str, Dict[date, str]],
) -> Dict[Tuple[date, str], Set[str]]:
    """
    Return a mapping of (shift_date, shift_name) -> set of display_names.

    Placement rule (duration-aware):
      - If a coach has ended-event duration data for a day, they are placed in
        the shift where they spent the most total session_duration seconds.
      - If a coach has NO ended-event data for a day (only untracked started
        events), fall back to whichever shift their started event was in.

    Fill-in rule (applied on top of placement):
      - If the resolved placement shift matches the coach's assigned shift,
        they appear only in that column.
      - If the resolved placement shift differs from the assigned shift, they
        appear in that column marked "(fill-in)".
    """
    # Build reverse lookups keyed on the fake DB coach card name (lower-case):
    #   fake_name -> assigned_shift
    #   fake_name -> real name
    fake_to_shift: Dict[str, str] = {}
    fake_to_real: Dict[str, str] = {}
    for _email, (shift, fake_name, real_name) in COACH_MAP.items():
        fake_to_shift[fake_name.lower()] = shift
        fake_to_real[fake_name.lower()] = real_name

    # Collect shifts seen from started events as fallback:
    # real_name -> shift_date -> set of shift_names (from started events)
    started_day_shifts: Dict[str, Dict[date, Set[str]]] = defaultdict(lambda: defaultdict(set))

    for event in events:
        if event.get("event_name") != "coach_auth_session_started":
            continue

        props = event.get("event_properties") or {}
        coach_name_raw = (props.get("coach_name") or "").strip()
        if not coach_name_raw or coach_name_raw.lower() == "test coach":
            continue

        created_at = event.get("created_at")
        if not created_at or not isinstance(created_at, datetime):
            continue

        ist_dt = _to_ist(created_at)
        shift_name, shift_date = _shift_for_ist_time(ist_dt)

        if shift_date < RANGE_START or shift_date > RANGE_END:
            continue

        real_name = fake_to_real.get(coach_name_raw.lower(), coach_name_raw)
        started_day_shifts[real_name][shift_date].add(shift_name)

    # Collect all (real_name, shift_date) pairs seen across both event types
    all_coach_days: Dict[str, Set[date]] = defaultdict(set)
    for real_name, day_map in started_day_shifts.items():
        all_coach_days[real_name].update(day_map.keys())
    for real_name, day_map in most_active_shift.items():
        all_coach_days[real_name].update(day_map.keys())

    # Build the result grid
    grid: Dict[Tuple[date, str], Set[str]] = defaultdict(set)

    for real_name, days in all_coach_days.items():
        assigned_shift = next(
            (fake_to_shift[fn] for fn, rn in fake_to_real.items() if rn == real_name),
            None,
        )

        for shift_date in days:
            # Duration-based placement takes priority; fall back to started-event shifts
            duration_shift = most_active_shift.get(real_name, {}).get(shift_date)
            if duration_shift:
                resolved_shift = duration_shift
            else:
                fallback_shifts = started_day_shifts.get(real_name, {}).get(shift_date, set())
                if not fallback_shifts:
                    continue
                # Among fallback shifts, prefer the assigned shift if present
                if assigned_shift and assigned_shift in fallback_shifts:
                    resolved_shift = assigned_shift
                else:
                    resolved_shift = next(iter(fallback_shifts))

            in_own_shift = (assigned_shift is not None) and (resolved_shift == assigned_shift)

            if in_own_shift:
                grid[(shift_date, assigned_shift)].add(real_name)
            else:
                grid[(shift_date, resolved_shift)].add(f"{real_name} (fill-in)")

    return grid


def build_excel(grid: Dict[Tuple[date, str], Set[str]]) -> str:
    """Write the shift activity grid to an Excel file and return the filename."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Coach Shift Activity"

    # ---- Styles ----
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    morning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # soft yellow
    evening_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # soft green
    night_fill   = PatternFill(start_color="DAE3F3", end_color="DAE3F3", fill_type="solid")  # soft blue

    shift_fills = {
        SHIFT_MORNING: morning_fill,
        SHIFT_EVENING: evening_fill,
        SHIFT_NIGHT:   night_fill,
    }

    # ---- Header row ----
    headers = ["Date", "Morning (5am–1pm)", "Evening (1pm–9pm)", "Night (9pm–5am)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 22

    # ---- Data rows ----
    current = RANGE_START
    row_idx = 2
    while current <= RANGE_END:
        date_label = current.strftime("%d %b %Y (%A)")

        def cell_text(shift: str) -> str:
            names = sorted(grid.get((current, shift), set()))
            return "\n".join(names) if names else ""

        row_data = [
            date_label,
            cell_text(SHIFT_MORNING),
            cell_text(SHIFT_EVENING),
            cell_text(SHIFT_NIGHT),
        ]
        ws.append(row_data)

        # Style each cell
        date_cell = ws.cell(row=row_idx, column=1)
        date_cell.font = Font(bold=True)
        date_cell.alignment = Alignment(horizontal="center", vertical="top")

        for col_offset, shift in enumerate(SHIFTS, start=2):
            cell = ws.cell(row=row_idx, column=col_offset)
            cell.fill = shift_fills[shift]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Highlight fill-in entries in italic
            text = cell.value or ""
            if "(fill-in)" in text:
                cell.font = Font(italic=True)

        # Auto-height: count newlines in the tallest cell
        max_lines = max(
            (v.count("\n") + 1 if v else 1)
            for v in row_data[1:]
        )
        ws.row_dimensions[row_idx].height = max(18, max_lines * 16)

        current += timedelta(days=1)
        row_idx += 1

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 32

    # ---- Freeze header ----
    ws.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename = f"coach_shift_activity_{timestamp}.xlsx"
    wb.save(filename)
    return filename


async def main():
    logger.info("Starting coach shift activity export...")
    events = await fetch_events()

    # Build fake->real name lookup once, shared by both helpers
    fake_to_real: Dict[str, str] = {
        fake_name.lower(): real_name
        for _email, (_shift, fake_name, real_name) in COACH_MAP.items()
    }

    most_active = compute_most_active_shift(events, fake_to_real)
    grid = bucket_events(events, most_active)
    filename = build_excel(grid)
    print(f"\n{'='*60}")
    print(f"Export completed successfully!")
    print(f"File: {filename}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    asyncio.run(main())
